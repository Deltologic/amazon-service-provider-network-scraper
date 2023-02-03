import time

from selenium import webdriver
import selenium
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
from furl import furl
from typing import List
from openpyxl import Workbook
from openpyxl import load_workbook
import pyotp
from const import (
    ACCOUNT_EMAIL,
    ACCOUNT_PASSWORD,
    TOKEN_2FA,

)


class Automation:
    def __init__(self):
        self.chrome_options = Options()
        self.chrome_options.add_argument("--window-size=1920,1080")
        self.chrome_options.add_argument("--disable-extensions")
        self.chrome_options.add_argument("--proxy-server='direct://'")
        self.chrome_options.add_argument("--proxy-bypass-list=*")
        self.chrome_options.add_argument("--start-maximized")
        self.chrome_options.add_argument("--headless")
        self.chrome_options.add_argument("--disable-gpu")
        self.chrome_options.add_argument("--disable-dev-shm-usage")
        self.chrome_options.add_argument("--no-sandbox")
        self.chrome_options.add_argument("--ignore-certificate-errors")
        self.amazon_service_network_link = 'https://sellercentral.amazon.com/tsba'
        self.totp_2fa_token = TOKEN_2FA
        self.account_email = ACCOUNT_EMAIL
        self.account_password = ACCOUNT_PASSWORD

        # For now manually links instead of parameters as this is easier.
        self.services_to_scrape = [
            {'category': 'Account Management', 'country': 'USA',
             'link': "https://sellercentral.amazon.com/tsba/searchpage/Account%20Management?ref_=sc_spn_hp_blst&sellFrom=US&sellIn=US"},
            {'category': 'Account Management', 'country': 'United Kingdom',
             'link': "https://sellercentral.amazon.com/tsba/searchpage/Account%20Management?ref_=sc_spn_hp_blst&sellFrom=UK&sellIn=UK"},
            {'category': 'Account Management', 'country': 'Germany',
             'link': "https://sellercentral.amazon.com/tsba/searchpage/Account%20Management?ref_=sc_spn_hp_blst&sellFrom=DE&sellIn=DE"},
            {'category': 'Accounting', 'country': 'USA',
             'link': "https://sellercentral.amazon.com/tsba/searchpage/Accounting?ref_=sc_spn_hp_acclst&sellFrom=US&sellIn=US"},
            {'category': 'Accounting', 'country': 'United Kingdom',
             'link': "https://sellercentral.amazon.com/tsba/searchpage/Accounting?ref_=sc_spn_hp_acclst&sellFrom=UK&sellIn=UK"},
            {'category': 'Accounting', 'country': 'Germany',
             'link': "https://sellercentral.amazon.com/tsba/searchpage/Accounting?ref_=sc_spn_hp_acclst&sellFrom=DE&sellIn=DE"},
            {'category': 'Advertising Optimization', 'country': 'USA',
             'link': "https://sellercentral.amazon.com/tsba/searchpage/Advertising%20Optimization?ref_=sc_spn_hp_alst&sellFrom=US&sellIn=US"},
            {'category': 'Advertising Optimization', 'country': 'United Kingdom',
             'link': "https://sellercentral.amazon.com/tsba/searchpage/Advertising%20Optimization?ref_=sc_spn_hp_alst&sellFrom=UK&sellIn=UK"},
            {'category': 'Advertising Optimization', 'country': 'Germany',
             'link': "https://sellercentral.amazon.com/tsba/searchpage/Advertising%20Optimization?ref_=sc_spn_hp_alst&sellFrom=DE&sellIn=DE"},
        ]
        self.driver = webdriver.Chrome(
            service=ChromeService(ChromeDriverManager().install()),
            options=self.chrome_options,
        )

    def scrape_services(self, service_dict: dict) -> List[dict]:
        self.driver.get(service_dict['link'])
        time.sleep(2)
        services = self.scrape_services_from_single_page(service_dict)
        print(services)
        while self.go_on_next_page():
            new_batch = self.scrape_services_from_single_page(service_dict)
            print(new_batch)
            services += new_batch
        return services

    def go_on_next_page(self) -> bool:

        try:
            self.driver.find_element(By.XPATH, "//li[@class='a-disabled a-last']")
            return False
        except NoSuchElementException:
            try:
                if self.driver.find_element(By.CLASS_NAME, 'a-last'):
                    self.driver.find_element(By.CLASS_NAME, 'a-last').click()
                    time.sleep(3)
                    return True
            except NoSuchElementException:
                return False
        return False

    def scrape_services_from_single_page(self, service_dict: dict) -> List[dict]:
        WebDriverWait(self.driver, 13).until(
            EC.presence_of_element_located((By.XPATH, "//div[@class='a-row provider-card-border']"))
        )
        html = self.driver.page_source
        soup = BeautifulSoup(html, features="html.parser")
        services_section = soup.find('div', {'class': 'searchResultsRightDiv'})
        services_list_html = services_section.find_all('div', {'class': 'a-row provider-card-border'})

        services_list = []

        for service in services_list_html:
            service_name = service.find('div', {'class': 'providerNamePadding'}).text
            description = service.find('div', {'class': 'specialitySection'}).text
            short_description = description[:len(description) // 2]
            short_description = short_description.replace('Specialities:', '')

            services_list.append(
                {'company_name': service_name,
                 'short_description': short_description,
                 'category': service_dict['category'],
                 'country': service_dict['country']}
            )

        return services_list

    def run(self):

        self.driver.get(self.amazon_service_network_link)
        time.sleep(5)
        self.login_to_amazon()
        filename = "amazon_service_network.xlsx"

        for service_dict in self.services_to_scrape:
            print(f"Scraping {service_dict['country']} - {service_dict['category']}")
            new_rows = self.scrape_services(service_dict)

            try:
                wb = load_workbook(filename)
                ws = wb.worksheets[0]  # select first worksheet
            except FileNotFoundError:
                headers_row = ['company_name', 'short_description', 'country', 'category']
                wb = Workbook()
                ws = wb.active
                ws.append(headers_row)

            field_names = ['company_name', 'short_description', 'country', 'category']

            for row in new_rows:
                values = (row[k] for k in field_names)
                ws.append(values)
                wb.save(filename)
            time.sleep(5)

    def login_to_amazon(self) -> bool:
        self.driver.get(self.amazon_service_network_link)
        email_element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.ID, "ap_email"))
        )
        email_element.send_keys(self.account_email)
        password_element = self.driver.find_element(By.ID, "ap_password")
        password_element.send_keys(self.account_password)
        time.sleep(1)
        submit_element = self.driver.find_element(By.ID, "signInSubmit")
        submit_element.submit()
        time.sleep(3)
        totp = pyotp.TOTP(self.totp_2fa_token)

        otp_element = WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located((By.ID, "auth-mfa-otpcode"))
        )
        otp_element.send_keys(totp.now())
        submit_element = self.driver.find_element(By.ID, "auth-signin-button")
        submit_element.submit()
        time.sleep(2)

        try:
            failed_otp = self.driver.find_element(By.ID, "auth-signin-button")
            otp_element = self.driver.find_element(By.ID, "auth-mfa-otpcode")
            otp_element.send_keys(totp.now())
            submit_element = self.driver.find_element(By.ID, "auth-signin-button")
            submit_element.submit()
            time.sleep(2)
            return True
        except NoSuchElementException:
            print("OAUTH worked...")
            return True


if __name__ == '__main__':
    print("Start")
    x = Automation()
    x.run()
    print("End")